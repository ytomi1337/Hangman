from openpyxl import  Workbook, load_workbook
from datetime import datetime
import random
import string
import msvcrt
import os

excel_file = load_workbook('words.xlsx')
excel_sheet = excel_file.active

excel_history_file = load_workbook('game_history.xlsx')
excel_history_sheet = excel_history_file.active

def word_roll():
    words_list = []
    for cell in excel_sheet['A']:
        words_list.append(cell.value)

    random_word = random.choice(words_list)

    return str(random_word).upper()

def hangman_quess():
    username = input("Podaj nazwe gracza: ")
    word = word_roll()
    word_letters = list(word)
    alphabet = set(string.ascii_uppercase)
    used_letter = set()
    lifes =['<3', '<3', '<3', '<3', '<3', '<3']
    attemps = 1
    data = datetime.now()
    format_date = data.strftime("%d.%m.%Y")

    for cell in excel_sheet['A']:
        if cell.value == word:
            nr_cell = str(cell.coordinate)
            nr_cell = nr_cell.upper().replace("A", "B")

            category = excel_sheet[nr_cell].value


    while len(word_letters) > 0:
        
        if lifes == []:
            print("Koniec gry ! Przegrana, słowo: ", word)
            print("")
            print("Kliknij dowolny przycisk by wrocic do menu....")
            msvcrt.getch()
            os.system('cls')

            excel_history_sheet.append([username, format_date, word, len(lifes), attemps, 'Przegrana'])
            excel_history_file.save('game_history.xlsx')
            return

        else:
            print("")
            print('Kategoria:',category,'Zycia:',' '.join(lifes))
            print('Zużyte litery: ', ' '.join(used_letter))
            word_blank = [letter if letter in used_letter else '-' for letter in word]
            print('Słowo: ', ' '.join(word_blank))
            print("")
            answer = input("Zgadnij wyraz: ").upper()

            attemps += 1

            if answer in alphabet - used_letter:
                used_letter.add(answer)
                if answer in word_letters:
                    os.system('cls')
                    for i,x in enumerate(word_letters):
                        if x == answer:
                            word_letters.remove(answer)
             
                else:
                    os.system('cls')
                    lifes.remove('<3')
            

            elif answer in used_letter:
                print("")
                print(f'Literka: {answer} Została już wprowadzona!')
                print("")
            else:
                print('Nie prawidłowa wartość!')
                attemps -= 1
    
    print(f'Gratulacje zgadłeś! Słowo: {word} Liczba prób: {attemps}')
    print("")
    print("Kliknij dowolny przycisk by wrocic do menu....")
    msvcrt.getch()
    os.system('cls')
    

    excel_history_sheet.append([username, format_date, word, len(lifes), attemps, 'Wygrana'])
    excel_history_file.save('game_history.xlsx')
            

def history():
    for row in excel_history_sheet.iter_rows(2, excel_history_sheet.max_row, values_only=True):
        for value in row:
            print(value, end='  |  ')
        print()
    
    print("")
    print("Kliknij dowolny przycisk by wrocic do menu....")
    print("")
    msvcrt.getch()
    os.system('cls')

def add_word():
    new_word = str(input("Podaj słowo: ")).upper()
    new_category = str(input("Podaj Kategorie: ")).upper()
    
    
    excel_sheet.append([new_word, new_category])
    excel_file.save('words.xlsx')

    print("")
    print("Dodano Poyślnie!")
    print("")
    print("Kliknij dowolny przycisk by wrocic do menu....")
    print("")
    msvcrt.getch()
    os.system('cls')

def main():
    choice_list = ['1', '2', '3', '4']

    while True:
        print("1: Zagraj")
        print("2: Dodaj Słowo")
        print("3: Historia Gier")
        print("4: Wyjdź")
        print("")
        choice = input("Wybór: ")

        if choice == '1':
            os.system('cls')
            hangman_quess()
        elif choice == '2':
            os.system('cls')
            add_word()
        elif choice == '3':
            os.system('cls')
            print("GRACZ  |  DATA ROZGRYWKI  |  SŁOWO  |  ZYC  |  PROB  |  WYNIK  |")
            history()
        elif choice == '4':
            os.system('cls')
            print("Opuszczanie gry")
            print("")
            print("Kliknij dowolny przycisk by wrocic do menu....")
            print("")
            msvcrt.getch()
            return
        elif choice is not choice_list:
            print("Nie prawidłowy wybór ! Wpisz ponownie")
            print("")

main()
